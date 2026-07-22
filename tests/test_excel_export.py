"""Exportación Excel del día.

Regresión de A1: el descuento del Excel usa totales.monto_descuento (clampeado
al subtotal), así la columna "Descuento" nunca supera al subtotal ni descuadra
contra "Total".
"""
from openpyxl import load_workbook

from app.database import SessionLocal
from app.excel_export import exportar_dia

FECHA = "2030-07-01"
# Columnas 1-based (ver COLUMNAS en excel_export): Descuento=8, Total=9.
COL_DESC = 8
COL_TOTAL = 9
PRIMERA_FILA = 3  # fila 1 título, fila 2 encabezados


def _crear(client, **extra):
    body = {"fecha": FECHA, "cliente_nombre": "Test", "tipo": "Take away", **extra}
    r = client.post("/api/pedidos", json=body)
    assert r.status_code == 200, r.text
    return r.json()


def test_descuento_export_clampeado_al_subtotal(client):
    # Subtotal 10000, descuento por monto absurdo (99999): debe clampearse.
    p = _crear(
        client,
        items=[{"nombre": "Milanesa", "cantidad": 1, "precio_unitario": 10000}],
        descuento_tipo="monto",
        descuento_valor=99999,
    )
    assert p["total"] == 0  # 10000 - 10000

    from datetime import date

    db = SessionLocal()
    try:
        ruta = exportar_dia(db, date.fromisoformat(FECHA))
    finally:
        db.close()

    assert ruta.exists()
    ws = load_workbook(ruta).active
    desc = ws.cell(row=PRIMERA_FILA, column=COL_DESC).value
    total = ws.cell(row=PRIMERA_FILA, column=COL_TOTAL).value
    # Clampeado al subtotal (10000), NO el valor crudo (99999).
    assert desc == 10000
    assert total == 0
