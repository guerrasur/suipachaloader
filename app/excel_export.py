"""Exportación a Excel mensual con openpyxl.

- Un archivo por mes: ``Suipacha_Pedidos_-_<Mes>_<Año>.xlsx``.
- Una hoja por día, nombrada ``Suipacha- DDMM``.
- Si la hoja del día ya existe se **regenera completa** desde la BD (se borra
  y se reescribe), nunca se duplica ni se parchea.
- Los pedidos anulados aparecen marcados como ``ANULADO`` y no suman.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .database import BASE_DIR
from .models import Pedido
from .totales import monto_envio

EXPORT_DIR = BASE_DIR / "exports"

MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
    "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

COLUMNAS = [
    ("N°", 6), ("Hora", 8), ("Tipo", 11), ("Cliente", 20), ("Dirección", 24),
    ("Ítems", 40), ("Envío", 10), ("Descuento", 11), ("Total", 12),
    ("Pago", 14), ("Detalle pago", 14), ("Repartidor", 14),
    ("Hora salida", 12), ("Facturado", 11), ("Notas", 20),
]

# Paleta alineada con la planilla de referencia (efectivo verde, transferencia
# azul) para que el Excel se lea de un vistazo.
_BRAND = "2F5496"
_TITLE_FILL = PatternFill("solid", fgColor=_BRAND)
_TITLE_FONT = Font(bold=True, size=15, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor=_BRAND)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SUM_FONT = Font(bold=True)
_SUM_TOTAL_FONT = Font(bold=True, size=12, color=_BRAND)
_CLIENTE_FONT = Font(bold=True)
_ANULADO_FONT = Font(color="9C9C9C", italic=True, strike=True)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = '$#,##0'

# Bandas para diferenciar filas (cebra).
_ZEBRA_FILL = PatternFill("solid", fgColor="F2F5FA")

# Rellenos + fuentes por método de pago (celda "Pago").
_PAGO_ESTILO = {
    "Efectivo":      (PatternFill("solid", fgColor="1F8A4C"), Font(bold=True, color="FFFFFF")),
    "Transferencia": (PatternFill("solid", fgColor="2F6FD0"), Font(bold=True, color="FFFFFF")),
    "QR":            (PatternFill("solid", fgColor="7A4FD0"), Font(bold=True, color="FFFFFF")),
    "Posnet":        (PatternFill("solid", fgColor="B26A00"), Font(bold=True, color="FFFFFF")),
}

# Estado de facturado.
_FACT_SI = (PatternFill("solid", fgColor="D6F0DE"), Font(bold=True, color="1F8A4C"))
_FACT_NO = (PatternFill("solid", fgColor="FDE3E0"), Font(bold=True, color="C0392B"))


def nombre_archivo(anio: int, mes: int) -> str:
    return f"Suipacha_Pedidos_-_{MESES[mes]}_{anio}.xlsx"


def nombre_hoja(d: date) -> str:
    return f"Suipacha- {d.strftime('%d%m')}"


def nombre_archivo_dia(d: date) -> str:
    return f"Suipacha_{d.strftime('%d-%m-%Y')}.xlsx"


def _items_texto(p: Pedido) -> str:
    return "\n".join(f"{i.cantidad}x {i.nombre}" for i in p.items)


def exportar_mes(db: Session, anio: int, mes: int) -> Path:
    """Genera/actualiza el archivo del mes. Devuelve la ruta al .xlsx."""
    EXPORT_DIR.mkdir(exist_ok=True)
    ruta = EXPORT_DIR / nombre_archivo(anio, mes)

    if ruta.exists():
        wb = load_workbook(ruta)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # sacamos la hoja vacía por defecto

    # Días del mes con pedidos, ordenados.
    pedidos = (
        db.query(Pedido)
        .filter(Pedido.fecha >= date(anio, mes, 1))
        .filter(Pedido.fecha < (date(anio + (mes == 12), (mes % 12) + 1, 1)))
        .order_by(Pedido.fecha, Pedido.hora_pedido)
        .all()
    )
    por_dia: dict[date, list[Pedido]] = {}
    for p in pedidos:
        por_dia.setdefault(p.fecha, []).append(p)

    for d in sorted(por_dia):
        _regenerar_hoja(wb, d, por_dia[d])

    # Ordenar las hojas por fecha para que queden prolijas.
    wb._sheets.sort(key=lambda ws: ws.title)
    wb.save(ruta)
    return ruta


def exportar_dia(db: Session, d: date) -> Path:
    """Genera un archivo con SOLO la hoja de ese día (misma hoja que en el
    mensual, ``Suipacha- DDMM``). Sirve para cargar/pegar esa hoja al final
    del día dentro del Excel del mes que junta todos los días."""
    EXPORT_DIR.mkdir(exist_ok=True)
    ruta = EXPORT_DIR / nombre_archivo_dia(d)

    pedidos = (
        db.query(Pedido)
        .filter(Pedido.fecha == d)
        .order_by(Pedido.hora_pedido)
        .all()
    )

    wb = Workbook()
    wb.remove(wb.active)  # sacamos la hoja vacía por defecto
    _regenerar_hoja(wb, d, pedidos)
    wb.save(ruta)
    return ruta


def _regenerar_hoja(wb: Workbook, d: date, pedidos: list[Pedido]) -> None:
    titulo = nombre_hoja(d)
    if titulo in wb.sheetnames:
        del wb[titulo]  # borrar y reescribir: nunca parchear
    ws = wb.create_sheet(titulo)

    # Anchos de columna.
    for idx, (_, ancho) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    # Título (banda de color, texto blanco centrado — como la planilla).
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
    tcell = ws.cell(row=1, column=1, value=f"Suipacha — Pedidos {d.strftime('%d/%m/%Y')}")
    tcell.font = _TITLE_FONT
    tcell.fill = _TITLE_FILL
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Encabezados.
    hrow = 2
    ws.row_dimensions[hrow].height = 20
    for idx, (nombre, _) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=hrow, column=idx, value=nombre)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    # Filas de pedidos.
    r = hrow + 1
    tot_dia = 0.0
    por_metodo: dict[str, float] = {}
    n_validos = 0
    banda = False  # alterna para la cebra
    for p in pedidos:
        hora = p.hora_pedido.strftime("%H:%M") if p.hora_pedido else ""
        salida = p.hora_salida.strftime("%H:%M") if p.hora_salida else ""
        envio = monto_envio(p)
        desc = _monto_descuento_export(p)
        fila = [
            p.numero or "", hora, p.tipo, p.cliente_nombre, p.cliente_direccion,
            _items_texto(p), envio or "", desc or "", p.total,
            p.metodo_pago, p.pago_efectivo_detalle, p.repartidor,
            salida, "Sí" if p.facturado else "No",
            ("ANULADO — " + p.notas) if p.anulado else p.notas,
        ]
        for idx, val in enumerate(fila, start=1):
            c = ws.cell(row=r, column=idx, value=val)
            c.border = _BORDER
            c.alignment = Alignment(
                vertical="top",
                wrap_text=(idx in (6, 15)),
                horizontal="center" if idx in (1, 10, 14) else None,
            )
            if idx in (7, 8, 9):  # envío / descuento / total
                c.number_format = _MONEY
            if p.anulado:
                # Anulado: gris tachado, sin colores de estado que distraigan.
                c.font = _ANULADO_FONT
                if banda:
                    c.fill = _ZEBRA_FILL
                continue
            if banda:
                c.fill = _ZEBRA_FILL
            if idx == 1:  # número de pedido en negrita
                c.font = _CLIENTE_FONT
            elif idx == 4:  # cliente en negrita
                c.font = _CLIENTE_FONT
            elif idx == 10:  # método de pago con color
                estilo = _PAGO_ESTILO.get(p.metodo_pago)
                if estilo:
                    c.fill, c.font = estilo
            elif idx == 14:  # facturado sí/no con color
                c.fill, c.font = _FACT_SI if p.facturado else _FACT_NO
        if not p.anulado:
            tot_dia += p.total
            por_metodo[p.metodo_pago] = por_metodo.get(p.metodo_pago, 0.0) + p.total
            n_validos += 1
        banda = not banda
        r += 1

    # Sumarios al pie.
    r += 1
    _sumario(ws, r, "Cantidad de pedidos", n_validos, money=False)
    r += 1
    _sumario(ws, r, "TOTAL DEL DÍA", tot_dia, destacado=True)
    for metodo in ("Efectivo", "Transferencia", "QR", "Posnet"):
        r += 1
        estilo = _PAGO_ESTILO.get(metodo)
        _sumario(ws, r, f"  {metodo}", por_metodo.get(metodo, 0.0),
                 fill=estilo[0] if estilo else None)

    ws.freeze_panes = "A3"


def _sumario(ws, row, etiqueta, valor, money=True, destacado=False, fill=None):
    ec = ws.cell(row=row, column=1, value=etiqueta)
    vc = ws.cell(row=row, column=9, value=valor)
    fuente = _SUM_TOTAL_FONT if destacado else _SUM_FONT
    ec.font = fuente
    vc.font = fuente
    if money:
        vc.number_format = _MONEY
    if fill is not None:
        # Chip de color en la etiqueta del método (texto blanco).
        ec.fill = fill
        ec.font = Font(bold=True, color="FFFFFF")


def _monto_descuento_export(p: Pedido) -> float:
    if not p.descuento_tipo or not p.descuento_valor:
        return 0.0
    sub = sum(i.cantidad * i.precio_unitario for i in p.items)
    if p.descuento_tipo == "porcentaje":
        return round(sub * p.descuento_valor / 100.0, 2)
    return float(p.descuento_valor)
